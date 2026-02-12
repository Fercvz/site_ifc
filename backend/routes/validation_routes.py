"""
Validation API Routes — Excel upload, validation results, and export endpoints.
"""

import io
import csv
from fastapi import APIRouter, UploadFile, File, Query, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook

from services.session_manager import session_manager
from services.excel_validator import extract_discipline_stage, parse_excel_rules, run_validation

router = APIRouter(prefix="/api", tags=["Validation"])


@router.post("/excel/upload")
async def upload_excel(
    file: UploadFile = File(...),
    session_id: str = Query(...),
):
    """Upload Excel spreadsheet and run validation against the loaded IFC."""
    session = session_manager.get_session(session_id)
    if not session or not session.get("ifc_index"):
        raise HTTPException(status_code=400, detail="Nenhum IFC processado nesta sessão. Carregue um IFC primeiro.")

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Apenas arquivos .xlsx são aceitos.")

    # Extract discipline and stage from IFC filename
    ifc_filename = session.get("ifc_filename", "")
    try:
        discipline, stage = extract_discipline_stage(ifc_filename)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # Parse Excel
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active

        # Read headers
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(str(cell.value).strip() if cell.value else "")

        # Validate required columns
        required = ["DISCIPLINA CATEGORIZADA", "CATEGORIA IFC", "Pset", "PROPRIEDADE IFC"]
        missing = [col for col in required if col not in headers]
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"Colunas obrigatórias ausentes na planilha: {', '.join(missing)}"
            )

        # Read rows
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = val
            rows.append(row_dict)

        wb.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao ler planilha: {str(e)}")

    # Parse rules
    try:
        rules = parse_excel_rules(rows, discipline, stage)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    if not rules:
        raise HTTPException(
            status_code=400,
            detail=f"Nenhuma regra encontrada para disciplina '{discipline}' e etapa '{stage}'."
        )

    # Run validation
    elements = session["ifc_index"].get("elements", [])
    results = run_validation(elements, rules)

    # Store results, include metadata
    results["discipline"] = discipline
    results["stage"] = stage
    results["ifc_filename"] = ifc_filename
    results["excel_filename"] = file.filename
    results["rules_count"] = len(rules)

    session_manager.update_session(session_id, validation_results=results)

    return {
        "status": "done",
        "discipline": discipline,
        "stage": stage,
        "summary": results["summary"],
        "rules_count": len(rules),
    }


@router.get("/validation/summary")
async def get_validation_summary(session_id: str = Query(...)):
    """Get global validation summary."""
    session = session_manager.get_session(session_id)
    if not session or not session.get("validation_results"):
        raise HTTPException(status_code=400, detail="Nenhuma validação executada nesta sessão.")

    vr = session["validation_results"]
    return {
        "summary": vr["summary"],
        "discipline": vr.get("discipline"),
        "stage": vr.get("stage"),
        "ifc_filename": vr.get("ifc_filename"),
        "excel_filename": vr.get("excel_filename"),
    }


@router.get("/validation/by-entity")
async def get_validation_by_entity(session_id: str = Query(...)):
    """Get validation results grouped by IFC entity type."""
    session = session_manager.get_session(session_id)
    if not session or not session.get("validation_results"):
        raise HTTPException(status_code=400, detail="Nenhuma validação executada nesta sessão.")

    return session["validation_results"]["by_entity"]


@router.get("/validation/by-property")
async def get_validation_by_property(session_id: str = Query(...)):
    """Get validation results grouped by property."""
    session = session_manager.get_session(session_id)
    if not session or not session.get("validation_results"):
        raise HTTPException(status_code=400, detail="Nenhuma validação executada nesta sessão.")

    return session["validation_results"]["by_property"]


@router.get("/validation/issues")
async def get_validation_issues(
    session_id: str = Query(...),
    entity: str = Query(default=None),
    reason: str = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=200),
):
    """Get detailed non-conformity issues with optional filters and pagination."""
    session = session_manager.get_session(session_id)
    if not session or not session.get("validation_results"):
        raise HTTPException(status_code=400, detail="Nenhuma validação executada nesta sessão.")

    issues = session["validation_results"]["issues"]

    # Apply filters
    if entity:
        issues = [i for i in issues if i["entity_type"] == entity]
    if reason:
        issues = [i for i in issues if i["reason"] == reason]

    total = len(issues)
    start = (page - 1) * page_size
    end = start + page_size
    paginated = issues[start:end]

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size,
        "issues": paginated,
    }


@router.get("/validation/export.csv")
async def export_csv(session_id: str = Query(...)):
    """Export non-conformity issues as CSV."""
    session = session_manager.get_session(session_id)
    if not session or not session.get("validation_results"):
        raise HTTPException(status_code=400, detail="Nenhuma validação executada nesta sessão.")

    issues = session["validation_results"]["issues"]
    vr = session["validation_results"]

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=[
        "global_id", "step_id", "entity_type", "name", "pset", "property",
        "expected", "actual", "reason",
    ])
    writer.writeheader()
    for issue in issues:
        writer.writerow({
            "global_id": issue.get("global_id", ""),
            "step_id": issue.get("step_id", ""),
            "entity_type": issue.get("entity_type", ""),
            "name": issue.get("name", ""),
            "pset": issue.get("pset", ""),
            "property": issue.get("property", ""),
            "expected": issue.get("expected", ""),
            "actual": issue.get("actual", ""),
            "reason": issue.get("reason", ""),
        })

    output.seek(0)
    filename = f"nao_conformidades_{vr.get('discipline', '')}_{vr.get('stage', '')}.csv"
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/validation/export.xlsx")
async def export_xlsx(session_id: str = Query(...)):
    """Export non-conformity issues as Excel."""
    session = session_manager.get_session(session_id)
    if not session or not session.get("validation_results"):
        raise HTTPException(status_code=400, detail="Nenhuma validação executada nesta sessão.")

    issues = session["validation_results"]["issues"]
    vr = session["validation_results"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Não Conformidades"

    headers = ["GUID", "STEP ID", "Entidade", "Nome", "Pset", "Propriedade", "Expected", "Actual", "Motivo"]
    ws.append(headers)

    for issue in issues:
        ws.append([
            issue.get("global_id", ""),
            issue.get("step_id", ""),
            issue.get("entity_type", ""),
            issue.get("name", ""),
            issue.get("pset", ""),
            issue.get("property", ""),
            issue.get("expected", ""),
            issue.get("actual", ""),
            issue.get("reason", ""),
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"nao_conformidades_{vr.get('discipline', '')}_{vr.get('stage', '')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
